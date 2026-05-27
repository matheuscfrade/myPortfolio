import importlib.util
import io
import zipfile
from pathlib import Path

from openpyxl import load_workbook


MODULE_PATH = Path(__file__).resolve().parent / "servidor_admin.py"
spec = importlib.util.spec_from_file_location("servidor_admin", MODULE_PATH)
servidor_admin = importlib.util.module_from_spec(spec)
spec.loader.exec_module(servidor_admin)


def sample_pdf_paths(count=2):
    files = sorted(servidor_admin.EXPORT_DIR.rglob("*.pdf"))
    assert len(files) >= count, "A pasta Documentos precisa conter PDFs para testar exportação."
    return files[:count]


def relative_pdf(path):
    return path.relative_to(servidor_admin.EXPORT_DIR).as_posix()


def test_exportar_returns_zip_with_selected_files_and_excel_metadata():
    first, second = sample_pdf_paths(2)
    first_rel = relative_pdf(first)
    second_rel = relative_pdf(second)

    client = servidor_admin.app.test_client()
    response = client.post(
        "/api/exportar",
        json={
            "documentos": [
                {
                    "arquivo": first_rel,
                    "nome": "Documento A",
                    "assunto": "Assunto A",
                    "categoria": "Comissao",
                    "ano": 2026,
                    "data": "2026-01-02",
                    "numero": "101",
                },
                {
                    "arquivo": second_rel,
                    "nome": "Documento B",
                    "assunto": "Assunto B",
                    "categoria": "Outro",
                    "ano": 2025,
                    "data": "2025-03-04",
                    "numero": "202",
                },
            ]
        },
    )

    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    assert "portfolio-documentos-" in response.headers["Content-Disposition"]

    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert sorted(archive.namelist()) == sorted([
            first_rel,
            second_rel,
            "informacoes_documentos.xlsx",
        ])
        assert archive.read(first_rel) == first.read_bytes()
        assert archive.read(second_rel) == second.read_bytes()

        workbook = load_workbook(io.BytesIO(archive.read("informacoes_documentos.xlsx")))
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows == [
            ("Nome", "Assunto", "Categoria", "Data"),
            ("Documento A", "Assunto A", "Comissao", "02-01-2026"),
            ("Documento B", "Assunto B", "Outro", "04-03-2025"),
        ]


def test_exportar_keeps_legacy_arquivos_payload():
    first = sample_pdf_paths(1)[0]
    first_rel = relative_pdf(first)

    client = servidor_admin.app.test_client()
    response = client.post("/api/exportar", json={"arquivos": [first_rel]})

    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert sorted(archive.namelist()) == [
            first_rel,
            "informacoes_documentos.xlsx",
        ]
        assert archive.read(first_rel) == first.read_bytes()

        workbook = load_workbook(io.BytesIO(archive.read("informacoes_documentos.xlsx")))
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows == [
            ("Nome", "Assunto", "Categoria", "Data"),
            (None, None, None, None),
        ]


def test_exportar_rejects_unsafe_paths():
    client = servidor_admin.app.test_client()
    response = client.post("/api/exportar", json={"arquivos": ["../secrets.pdf"]})

    assert response.status_code == 400


def test_config_api_is_read_only_and_shared_config_uses_installer_file():
    client = servidor_admin.app.test_client()
    original = servidor_admin.load_app_config(servidor_admin.RUNTIME)

    response = client.post("/api/config", json={"displayName": "Outro Nome"})
    assert response.status_code == 405

    config = client.get("/api/config")
    assert config.status_code == 200
    assert config.json["displayName"] == original["displayName"]
    assert config.json["subtitle"] == original["subtitle"]

    shared = client.get("/shared-data/config.json")
    assert shared.status_code == 200
    assert shared.json["displayName"] == original["displayName"]


def test_public_package_api_creates_output_folder():
    client = servidor_admin.app.test_client()
    response = client.post("/api/gerar-pacote-publico")

    assert response.status_code == 200
    assert response.json["success"] is True
    assert "dist-publico" in response.json["path"]


def test_public_package_api_returns_json_error(monkeypatch=None):
    client = servidor_admin.app.test_client()
    original = servidor_admin.create_public_package_for_runtime

    def fail():
        raise PermissionError("sem permissao para escrever dist-publico")

    servidor_admin.create_public_package_for_runtime = fail
    try:
        response = client.post("/api/gerar-pacote-publico")
    finally:
        servidor_admin.create_public_package_for_runtime = original

    assert response.status_code == 500
    assert "sem permissao" in response.json["error"]


def test_local_public_portfolio_routes_redirect_to_admin():
    client = servidor_admin.app.test_client()

    root = client.get("/")
    portfolio = client.get("/portfolio/")
    portfolio_asset = client.get("/portfolio/index.html")

    assert root.status_code == 302
    assert root.headers["Location"] == "/admin/"
    assert portfolio.status_code == 302
    assert portfolio.headers["Location"] == "/admin/"
    assert portfolio_asset.status_code == 302
    assert portfolio_asset.headers["Location"] == "/admin/"


def test_sync_does_not_spawn_subprocess():
    sentinel = type("SubprocessSentinel", (), {})()

    def fail_if_called(*args, **kwargs):
        raise AssertionError("sync must not spawn sys.executable in the packaged app")

    sentinel.run = fail_if_called
    previous = getattr(servidor_admin, "subprocess", None)
    servidor_admin.subprocess = sentinel
    try:
        client = servidor_admin.app.test_client()
        response = client.post("/api/sync")
    finally:
        if previous is None:
            delattr(servidor_admin, "subprocess")
        else:
            servidor_admin.subprocess = previous

    assert response.status_code == 200
    assert response.json["success"] is True


if __name__ == "__main__":
    test_exportar_returns_zip_with_selected_files_and_excel_metadata()
    test_exportar_keeps_legacy_arquivos_payload()
    test_exportar_rejects_unsafe_paths()
    test_config_api_is_read_only_and_shared_config_uses_installer_file()
    test_public_package_api_creates_output_folder()
    test_public_package_api_returns_json_error()
    test_local_public_portfolio_routes_redirect_to_admin()
    test_sync_does_not_spawn_subprocess()
